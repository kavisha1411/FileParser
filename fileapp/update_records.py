import logging
import csv
import decimal

import pandas as pd
from datetime import datetime, date
from openpyxl import load_workbook

from .models import Upload, BookingData
from .settings import (
    MSG_TYPE_SUCCESS,
    MSG_TYPE_WARNING,
    MSG_TYPE_WARNING_LIST,
    ERROR_MSG_GENERIC,
    ERROR_MSG_DB_UPDATE,
    UPLOAD_DIR,
)

logger = logging.getLogger('fileapp')


def is_new_row(row, old_data, file_extension):
    if file_extension == 'xlsx':
        for old_row in old_data:
            if row[0].value == old_row.booking_id:
                return False
        return True

    else:
        for old_row in old_data:
            if row[0] == str(old_row.booking_id):
                return False
        return True


def data_is_modified(row, old_data, file_extension):
    fields = [field.name for field in BookingData._meta.fields]
    changed_fields = {}
    changed_booking_id = 0
    changed = 0

    if file_extension == 'xlsx':
        for old_row in old_data:
            booking_id = getattr(old_row, 'booking_id')
            if row[0].value == booking_id:
                for index in range(0, 26):
                    attribute = fields[index + 1]
                    value = getattr(old_row, attribute)
                    if row[index].value != value:
                        if (isinstance(row[index].value, int)) and (int(value) == row[index].value):
                            continue
                        elif (isinstance(row[index].value, str)) and (row[index].value == str(value)):
                            continue
                        elif (isinstance(row[index].value, datetime)) and (row[index].value.date() == value):
                            continue
                        else:
                            changed_booking_id = getattr(old_row, 'booking_id')
                            changed_fields[attribute] = row[index].value
                            changed = 1

    else:
        for old_row in old_data:
            booking_id = getattr(old_row, 'booking_id')
            if int(row[0]) == booking_id:
                for index in range(0, 26):
                    attribute = fields[index + 1]
                    value = getattr(old_row, attribute)
                    if row[index] != value:
                        if (isinstance(row[index], int)) and (int(value) == row[index]):
                            continue
                        elif isinstance(value, date):
                            valid_date = datetime.strptime(row[index], "%m/%d/%Y")
                            if valid_date.date() == value:
                                continue
                        elif isinstance(row[index], str):
                            if isinstance(value, int) and row[index] == str(value):
                                continue
                            elif isinstance(value, decimal.Decimal) and decimal.Decimal(row[index]) == value:
                                continue
                            else:
                                changed_booking_id = getattr(old_row, 'booking_id')
                                changed_fields[attribute] = row[index]
                                changed = 1

    return changed, changed_booking_id, changed_fields


def update_xlsx_data(uploaded_file):
    path = UPLOAD_DIR + f'{uploaded_file}'
    row_array = []
    file_extension = uploaded_file.name[-4:]
    workbook = load_workbook(path)
    xlsxfile = workbook.active
    count = 0

    try:
        old_data = BookingData.objects.all()
        for row in xlsxfile.iter_rows(min_row=2):
            if is_new_row(row, old_data, file_extension):
                count += 1
                row_array.append(BookingData(booking_id=row[0].value, code=row[1].value,
                                             origin=row[2].value, property=row[3].value,
                                             property_code=row[4].value, arrival=row[5].value,
                                             departure=row[6].value, accomodation=row[7].value,
                                             reservation_holder=row[8].value, reservation_total=row[9].value,
                                             tax=row[10].value, currency=row[11].value, payment_method=row[12].value,
                                             status=row[13].value, date_of_creation=row[14].value,
                                             email=row[15].value, phone=row[16].value, country=row[17].value,
                                             language=row[18].value, address=row[19].value, zipcode=row[20].value,
                                             city=row[21].value, access_code=row[22].value, b2b_discount=row[23].value,
                                             rooming_info=row[24].value, deposit=row[25].value,
                                             rate_plan_name=row[26].value))
            is_data_modified, booking_id, modified_fields = data_is_modified(
                row, old_data, file_extension)
            if is_data_modified:
                BookingData.objects.filter(booking_id=booking_id).update(**modified_fields)

        if row_array:
            BookingData.objects.bulk_create(row_array)
            msg = f'Total {count} record(s) created in database'
            msg_type = MSG_TYPE_SUCCESS

        else:
            count = BookingData.objects.all().count()
            msg = f'Total {count} record(s) in database'
            msg_type = MSG_TYPE_SUCCESS

    except Exception as e:
        msg = ERROR_MSG_DB_UPDATE
        msg_type = MSG_TYPE_WARNING

    return msg, msg_type


def update_csv_data(uploaded_file):
    path = UPLOAD_DIR + f'{uploaded_file}'
    row_array = []
    file_extension = uploaded_file.name[-4:]
    count = 0

    csvfile = open(path)
    csvreader = csv.reader(csvfile)
    next(csvreader)

    try:
        old_data = BookingData.objects.all()
        for row in csvreader:
            if is_new_row(row, old_data, file_extension):
                count += 1
                arrival = datetime.strptime(row[5], "%m/%d/%Y")
                departure = datetime.strptime(row[6], "%m/%d/%Y")
                date_creation = datetime.strptime(row[14], "%m/%d/%Y")
                row_array.append(BookingData(booking_id=row[0], code=row[1],
                                             origin=row[2], property=row[3],
                                             property_code=row[4], arrival=arrival,
                                             departure=departure, accomodation=row[7],
                                             reservation_holder=row[8], reservation_total=row[9],
                                             tax=row[10], currency=row[11], payment_method=row[12],
                                             status=row[13], date_of_creation=date_creation,
                                             email=row[15], phone=row[16], country=row[17],
                                             language=row[18], address=row[19], zipcode=row[20],
                                             city=row[21], access_code=row[22], b2b_discount=row[23],
                                             rooming_info=row[24], deposit=row[25],
                                             rate_plan_name=row[26]))
            is_data_modified, booking_id, modified_fields = data_is_modified(
                    row, old_data, file_extension)
            if is_data_modified:
                BookingData.objects.filter(booking_id=booking_id).update(**modified_fields)
        if row_array:
            BookingData.objects.bulk_create(row_array)
            msg = f'Total {count} record(s) created in database'
            msg_type = MSG_TYPE_SUCCESS

        else:
            count = BookingData.objects.all().count()
            msg = f'Total {count} record(s) in database'
            msg_type = MSG_TYPE_SUCCESS

    except Exception as e:
        msg = ERROR_MSG_DB_UPDATE
        msg_type = MSG_TYPE_WARNING

    return msg, msg_type
