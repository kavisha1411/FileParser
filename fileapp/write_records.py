import logging
import csv
import decimal
from django.db import transaction

import pandas as pd
from datetime import datetime, date
from openpyxl import load_workbook

from .models import Upload, BookingData
from .settings import (
    MSG_TYPE_SUCCESS,
    MSG_TYPE_WARNING,
    MSG_TYPE_WARNING_LIST,
    ERROR_MSG_GENERIC,
    ERROR_MSG_DB,
    UPLOAD_DIR,
)
from .utils.validation import validate_data

logger = logging.getLogger('fileapp')


def insert_xlsx_data(uploaded_file):
    path = UPLOAD_DIR + f'{uploaded_file}'
    BookingData.objects.all().delete()
    row_array = []
    workbook = load_workbook(path)
    xlsxfile = workbook.active

    try:
        for row in xlsxfile.iter_rows(min_row=2):
            row_array.append(BookingData(booking_id=row[0].value, code=row[1].value,
                                         origin=row[2].value, property=row[3].value,
                                         property_code=row[4].value, arrival=row[5].value,
                                         departure=row[6].value, accomodation=row[7].value,
                                         reservation_holder=row[8].value, reservation_total=row[9].value,
                                         tax=row[10].value, currency=row[11].value, payment_method=row[12].value,
                                         status=row[13].value, date_of_creation=row[14].value,
                                         email=row[15].value, phone=row[16].value, country=row[17].value,
                                         language=row[18].value, address=row[19].value, zipcode=row[20].value,
                                         city=row[21].value, access_code=row[22].value, b2b_discount=row[23].value,
                                         rooming_info=row[24].value, deposit=row[25].value,
                                         rate_plan_name=row[26].value))

        BookingData.objects.bulk_create(row_array)
        df = pd.DataFrame(list(BookingData.objects.all().values()))
        error_data = validate_data(df, 'normal')
        if not error_data:
            count = BookingData.objects.all().count()
            msg = f'Total {count} records in database'
            msg_type = MSG_TYPE_SUCCESS
        else:
            msg = error_data
            msg_type = MSG_TYPE_WARNING_LIST

    except Exception as e:
        msg = e
        msg_type = MSG_TYPE_WARNING

    return msg, msg_type


def insert_csv_data(uploaded_file):
    path = UPLOAD_DIR + f'{uploaded_file}'
    # BookingData.objects.all().delete()
    row_array = []

    csvfile = open(path)
    csvreader = csv.reader(csvfile)
    next(csvreader)
    try:
        for row in csvreader:
            arrival = datetime.strptime(row[5], "%m/%d/%Y")
            departure = datetime.strptime(row[6], "%m/%d/%Y")
            date_creation = datetime.strptime(row[14], "%m/%d/%Y")
            row_array.append(BookingData(booking_id=row[0], code=row[1],
                                         origin=row[2], property=row[3],
                                         property_code=row[4], arrival=arrival,
                                         departure=departure, accomodation=row[7],
                                         reservation_holder=row[8], reservation_total=row[9],
                                         tax=row[10], currency=row[11], payment_method=row[12],
                                         status=row[13], date_of_creation=date_creation,
                                         email=row[15], phone=row[16], country=row[17],
                                         language=row[18], address=row[19], zipcode=row[20],
                                         city=row[21], access_code=row[22], b2b_discount=row[23],
                                         rooming_info=row[24], deposit=row[25],
                                         rate_plan_name=row[26]))

        BookingData.objects.bulk_create(row_array)
        df = pd.DataFrame(list(BookingData.objects.all().values()))
        error_data = validate_data(df, 'normal')
        if not error_data:
            count = BookingData.objects.all().count()
            msg = f'Total {count} records in database'
            msg_type = MSG_TYPE_SUCCESS
        else:
            msg = error_data
            msg_type = MSG_TYPE_WARNING_LIST

    except Exception as e:
        msg = e
        msg_type = MSG_TYPE_WARNING

    return msg, msg_type


def insert_xlsx_data_pandas(uploaded_file):
    path = UPLOAD_DIR + f'{uploaded_file}'
    # BookingData.objects.all().delete()
    row_array = []
    df = pd.read_excel(path)

    try:
        for index, row in df.iterrows():
            row_array.append(BookingData(booking_id=row['ID'], code=row['Code'],
                                         origin=row['Origin'], property=row['Property'],
                                         property_code=row['Property - Code'], arrival=row['Arrival'],
                                         departure=row['Departure'], accomodation=row['Accommodation'],
                                         reservation_holder=row['Reservation holder'],
                                         reservation_total=row['Reservation total (EUR)'],
                                         tax=row['Tax'], currency=row['Currency'], payment_method=row['Payment method'],
                                         status=row['Status'], date_of_creation=row['Date of creation'],
                                         email=row['E-mail'], phone=row['Phone number'], country=row['Country'],
                                         language=row['Language'], address=row['Address'],
                                         zipcode=row['ZIP/Postal code'],
                                         city=row['City'], access_code=row['Access code'],
                                         b2b_discount=row['B2B discount applied'],
                                         rooming_info=row['Rooming info'], deposit=row['Deposit'],
                                         rate_plan_name=row['Rate plan name']))

        error_data = validate_data(df, 'pandas')
        if not error_data:
            BookingData.objects.bulk_create(row_array)
            index = df.index
            msg = f'Total {len(index)} records written'
            msg_type = MSG_TYPE_SUCCESS
        else:
            msg = error_data
            msg_type = MSG_TYPE_WARNING_LIST

    except Exception as e:
        msg = e
        msg_type = MSG_TYPE_WARNING

    return msg, msg_type


def insert_csv_data_pandas(uploaded_file):
    path = UPLOAD_DIR + f'{uploaded_file}'
    df = pd.read_csv(path)
    # BookingData.objects.all().delete()
    row_array = []

    try:
        for index, row in df.iterrows():
            arrival = datetime.strptime(row['Arrival'], "%m/%d/%Y")
            departure = datetime.strptime(row['Departure'], "%m/%d/%Y")
            date_creation = datetime.strptime(row['Date of creation'], "%m/%d/%Y")
            row_array.append(BookingData(booking_id=row['ID'], code=row['Code'],
                                         origin=row['Origin'], property=row['Property'],
                                         property_code=row['Property - Code'], arrival=arrival,
                                         departure=departure, accomodation=row['Accommodation'],
                                         reservation_holder=row['Reservation holder'],
                                         reservation_total=row['Reservation total (EUR)'],
                                         tax=row['Tax'], currency=row['Currency'],
                                         payment_method=row['Payment method'],
                                         status=row['Status'], date_of_creation=date_creation,
                                         email=row['E-mail'], phone=row['Phone number'], country=row['Country'],
                                         language=row['Language'], address=row['Address'],
                                         zipcode=row['ZIP/Postal code'],
                                         city=row['City'], access_code=row['Access code'],
                                         b2b_discount=row['B2B discount applied'],
                                         rooming_info=row['Rooming info'], deposit=row['Deposit'],
                                         rate_plan_name=row['Rate plan name']))

        error_data = validate_data(df, 'pandas')
        if not error_data:
            BookingData.objects.bulk_create(row_array)
            index = df.index
            msg = f'Total {len(index)} records written'
            msg_type = MSG_TYPE_SUCCESS
        else:
            msg = error_data
            msg_type = MSG_TYPE_WARNING_LIST

    except Exception as e:
        msg = e
        msg_type = MSG_TYPE_WARNING

    return msg, msg_type
